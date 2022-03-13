import time 
import openpyxl
from openpyxl import load_workbook
import xlrd

def FindTaskInPrgs():
    workbook = xlrd.open_workbook(r"C:\Users\andyl\OneDrive\Desktop\Gastops\Work Log.xlsx")
    sheet = workbook.sheet_by_name("Logger")
    rowcount = sheet.nrows
    colcount = sheet.ncols
    #print(rowcount)
    #print(colcount)
    InPrgsTask=[]
    InPrgsTaskrowcount=0
    for curr_row in range(1,rowcount):
        if sheet.cell_value(curr_row,2)=='In progress':
            row_data=[]
            row_data.append(curr_row)
            InPrgsTaskrowcount=InPrgsTaskrowcount+1
            for curr_col in range(0, colcount-1, 1):
                data = sheet.cell_value(curr_row, curr_col)
                #print(data)
                row_data.append(data)
            InPrgsTask.append(row_data)
    print('You have following tasks in progress:\nIndex , Task      ,     Start Time\n------------------------------------------')
    for IPTcount in range(0,InPrgsTaskrowcount):
        print(InPrgsTask[IPTcount])

def main():
    FindTaskInPrgs()
    wb=load_workbook(r"C:\Users\andyl\OneDrive\Desktop\Gastops\Work Log.xlsx")
    sheet=wb.active
    sheet=wb['Logger']
    for counts in range(0,100):
        IndORNewTask=input('__________________________________________\nEnter index of task you have finished:\n              OR              \nEnter New Task: (Enter * or space to exit program)\n__________________________________________\n\n')
        print('\n__________________________________________\n')
        if IndORNewTask=='*' or IndORNewTask==' ':
            print('******Exiting program')
            break
        if (IndORNewTask.isdigit()==True):
            
            sheet.cell(row=int(IndORNewTask)+1, column=3).value=time.ctime()
            wb.save(r"C:\Users\andyl\OneDrive\Desktop\Gastops\Work Log.xlsx")
            print('******Task #'+str(IndORNewTask)+' complete')
            FindTaskInPrgs()
        else:
            local_time=time.ctime()
            DataInput=[(IndORNewTask,local_time,'In progress')]
            for i in DataInput:
                sheet.append(i)
            wb.save(r"C:\Users\andyl\OneDrive\Desktop\Gastops\Work Log.xlsx")
            print('******New task added')
            FindTaskInPrgs()
    

main()

